
# -*- coding: utf-8 -*-
"""
Hanwha XLSM – in-memory odstranění ochran + export sloupců do .xlsx
- Upravené .xlsm NEukládá na disk (jen při -debug do temp).
- Při -debug: TEMP složka → rozbalí PŮVODNÍ .xlsm (obsah ZIPu), upravené .xlsm uloží jako soubor (bez rozbalení).
- Výstup: .xlsx s vybranými sloupci (vždy s hlavičkou).
- Logování: konzole + soubor, s timestamp; na konci oddělovač pro přehlednost.
- Exit kódy pro plánovač úloh (viz níže).
"""

import sys
import io
import re
import zipfile
import argparse
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Optional

from openpyxl import load_workbook
import pandas as pd

# ---- Exit codes (pro plánovač úloh) ----
EXIT_OK = 0
EXIT_INPUT_NOT_FOUND = 1
EXIT_INMEM_MOD_FAIL = 2
EXIT_SHEET_ERROR = 3
EXIT_COLSPEC_ERROR = 4
EXIT_WRITE_FAIL = 5
EXIT_UNKNOWN_ERROR = 6


# ====== Logging setup ======

def setup_logger(log_path: Path) -> logging.Logger:
    """
    Nastaví logger pro soubor i konzoli. Přidá timestamp.
    Na konci běhu je vhodné zapsat oddělovač.
    """
    logger = logging.getLogger("hanwha")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    # File handler
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    # Console handler
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    # Úvodní hlavička běhu
    logger.info("=== START běhu ===")
    return logger


def write_run_separator(logger: logging.Logger, log_path: Path):
    """
    Zapíše oddělovač na konec běhu pro přehlednost v logu.
    """
    sep = "—" * 80
    logger.info(sep)
    # Ujistíme se, že je to zapsané i do souboru (flush přes uzavření handlerů neřešíme — logger to obsluhuje)


# ====== 1) Odstranění ochran v paměti (bytes regex) ======

def remove_tag_bytes(xml_bytes: bytes, tag_local_name: str) -> bytes:
    """
    Odstraní self-closing element <...tag_local_name.../> na úrovni bytes.
    Příklady: 'sheetProtection', 'workbookProtection'.
    Zachová vše ostatní beze změny.
    """
    pattern = re.compile(
        rb'<[^>]*' + re.escape(tag_local_name.encode('utf-8')) + rb'[^>]*/>',
        re.IGNORECASE
    )
    return pattern.sub(b'', xml_bytes)


def build_modified_xlsm_bytes(input_xlsm: Path, sheet_xml_rel: str, logger: logging.Logger) -> Tuple[bytes, dict]:
    """
    Načte .xlsm do paměti, upraví:
      - zadaný sheet XML (odstraní sheetProtection),
      - xl/workbook.xml (odstraní workbookProtection),
    a vrátí bajty UPRAVENÉHO .xlsm + statistiky (bytes před/po).
    """
    try:
        original_bytes = input_xlsm.read_bytes()
    except Exception as e:
        logger.error(f"Nelze číst vstupní soubor: {e}")
        raise

    in_mem_zip = io.BytesIO(original_bytes)
    out_mem_zip = io.BytesIO()

    stats = {
        "sheet_xml_before": 0,
        "sheet_xml_after": 0,
        "workbook_before": 0,
        "workbook_after": 0
    }

    try:
        with zipfile.ZipFile(in_mem_zip, 'r') as zin, zipfile.ZipFile(out_mem_zip, 'w', zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)

                if info.filename == sheet_xml_rel:
                    stats["sheet_xml_before"] = len(data)
                    data = remove_tag_bytes(data, 'sheetProtection')
                    stats["sheet_xml_after"] = len(data)
                    logger.info(f"{sheet_xml_rel}: {stats['sheet_xml_before']} → {stats['sheet_xml_after']} bytes "
                                f"(sheetProtection odstraněn: {stats['sheet_xml_before'] != stats['sheet_xml_after']})")

                elif info.filename == 'xl/workbook.xml':
                    stats["workbook_before"] = len(data)
                    data = remove_tag_bytes(data, 'workbookProtection')
                    stats["workbook_after"] = len(data)
                    logger.info(f"xl/workbook.xml: {stats['workbook_before']} → {stats['workbook_after']} bytes "
                                f"(workbookProtection odstraněn: {stats['workbook_before'] != stats['workbook_after']})")

                zout.writestr(info, data)
    except Exception as e:
        logger.error(f"In-memory úprava XLSM selhala: {e}")
        raise

    return out_mem_zip.getvalue(), stats


# ====== 2) Debug režim: TEMP (rozbalit PŮVODNÍ, uložit UPRAVENÝ) ======

def debug_temp_layout(original_xlsm: Path, mod_bytes: bytes, base_name: str,
                      sheet_xml_rel: str, logger: logging.Logger) -> Path:
    """
    Vytvoří TEMP složku, rozbalí PŮVODNÍ .xlsm obsah a uloží UPRAVENÝ .xlsm jako soubor.
    UPRAVENÝ .xlsm se NEROZBALUJE.
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_dir = Path.cwd() / f"_tempHanwha_{ts}"
    try:
        temp_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        logger.error(f"Nelze vytvořit TEMP složku '{temp_dir}': {e}")
        raise

    # 2a) Rozbalit PŮVODNÍ XLSM
    try:
        with zipfile.ZipFile(original_xlsm, 'r') as zorig:
            zorig.extractall(temp_dir)
        logger.info(f"[DEBUG] Rozbalen PŮVODNÍ XLSM do: {temp_dir}")
        logger.info(f"        → Zkontroluj: {temp_dir / sheet_xml_rel}")
        logger.info(f"        → Zkontroluj: {temp_dir / 'xl' / 'workbook.xml'}")
    except Exception as e:
        logger.error(f"Nelze rozbalit PŮVODNÍ XLSM do TEMP: {e}")
        raise

    # 2b) Uložit UPRAVENÝ XLSM jen jako soubor
    mod_xlsm_path = temp_dir / f"{base_name}_modified_{ts}.xlsm"
    try:
        mod_xlsm_path.write_bytes(mod_bytes)
        logger.info(f"[DEBUG] UPRAVENÝ XLSM uložen: {mod_xlsm_path} (nerozbalen)")
    except Exception as e:
        logger.error(f"Nelze uložit UPRAVENÝ XLSM do TEMP: {e}")
        raise

    return temp_dir


# ====== 3) Načtení listu přímo z bajtů (bez disku) ======

def load_sheet_rows_from_bytes(xlsm_bytes: bytes, sheet_name: str, logger: logging.Logger) -> List[Tuple]:
    """
    Načte řádky z listu přímo z bajtů (BytesIO), bez ukládání upraveného .xlsm na disk.
    """
    try:
        bio = io.BytesIO(xlsm_bytes)
        wb = load_workbook(bio, data_only=True, keep_vba=True, read_only=False)
    except Exception as e:
        logger.error(f"Nelze otevřít upravené XLSM z paměti: {e}")
        raise

    if sheet_name not in wb.sheetnames:
        logger.error(f"List '{sheet_name}' nebyl nalezen. Dostupné: {wb.sheetnames}")
        raise ValueError("SHEET_NOT_FOUND")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        logger.error(f"List '{sheet_name}' neobsahuje žádná data.")
        raise ValueError("SHEET_EMPTY")
    return rows


# ====== 4) Rezoluce sloupců a hlaviček + DataFrame ======

def resolve_columns(rows: List[Tuple], col_spec: str, logger: logging.Logger) -> Tuple[List[int], List[str], List[Tuple]]:
    """
    Určí sloupce podle col_spec:
      - "0,3" => indexy
      - "Model,Qty" => názvy v hlavičce (první řádek)
    Vrací (selected_indices, column_names, data_rows_bez_hlavičky)
    """
    if not col_spec or not col_spec.strip():
        logger.error("Specifikace sloupců je prázdná.")
        raise ValueError("COLSPEC_EMPTY")

    header = rows[0]
    parts = [p.strip() for p in col_spec.split(",") if p.strip()]

    try:
        if all(part.isdigit() for part in parts):
            selected = [int(p) for p in parts]
            max_idx = len(header) - 1
            bad = [i for i in selected if i < 0 or i > max_idx]
            if bad:
                logger.error(f"Neplatné indexy sloupců: {bad}; max index je {max_idx}")
                raise IndexError("COLSPEC_BAD_INDEX")
        else:
            name_to_idx = {str(h).strip(): i for i, h in enumerate(header)}
            missing = [n for n in parts if n not in name_to_idx]
            if missing:
                logger.error(f"Nenalezeny hlavičky: {missing}")
                raise KeyError("COLSPEC_HEADER_MISSING")
            selected = [name_to_idx[n] for n in parts]
    except Exception as e:
        # Přepošleme dál, ale ať máme log
        logger.error(f"Chyba při zpracování specifikace sloupců: {e}")
        raise

    column_names: List[str] = []
    for i in selected:
        h = header[i] if i < len(header) else None
        column_names.append(str(h).strip() if h is not None else f"Col{i}")

    data_rows = rows[1:]  # hlavičku do dat nepoužijeme, je v columns
    return selected, column_names, data_rows


def build_filtered_dataframe(data_rows: List[Tuple], selected_indices: List[int], column_names: List[str]) -> pd.DataFrame:
    """
    Sestaví DataFrame ze zvolených sloupců a nastaví názvy sloupců.
    """
    filtered = [[r[i] if i < len(r) else None for i in selected_indices] for r in data_rows]
    return pd.DataFrame(filtered, columns=column_names)


# ====== 5) Výpočet výstupní cesty pro .xlsx ======

def compute_output_path(input_xlsm: Path, out_arg: Optional[str], sheet_name: str) -> Path:
    """
    Pokud je zadán --out, použije se přímo; jinak '<base>_<sheet>_filtered_<timestamp>.xlsx' vedle vstupu.
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if out_arg:
        return Path(out_arg).resolve()
    base = input_xlsm.stem
    out_name = f"{base}_{sheet_name.replace(' ', '')}_filtered_{ts}.xlsx"
    return input_xlsm.with_name(out_name).resolve()


# ====== 6) CLI a hlavní běh ======

def main():
    parser = argparse.ArgumentParser(
        description="Hanwha XLSM – in-memory ochrany + export sloupců (debug: PŮVODNÍ rozbalit, UPRAVENÝ uložit)"
    )
    parser.add_argument("input", help="Vstupní .xlsm soubor")
    parser.add_argument("--sheet-xml", default="xl/worksheets/sheet2.xml",
                        help="Relativní cesta k sheet XML v ZIPu (default: xl/worksheets/sheet2.xml)")
    parser.add_argument("--sheet-name", default="Stock Raw",
                        help="Název listu k exportu (default: Stock Raw)")
    parser.add_argument("--cols", default="0,3",
                        help="Sloupce k exportu: indexy ('0,3') nebo názvy hlaviček ('Model,Qty')")
    parser.add_argument("--out", default=None,
                        help="Cesta k výstupnímu .xlsx (implicitně '<base>_<sheet>_filtered_<timestamp>.xlsx')")
    parser.add_argument("--log", default="hanwha_inmem_export.log",
                        help="Cesta k log souboru (default: hanwha_inmem_export.log)")
    parser.add_argument("-debug", action="store_true",
                        help="TEMP: rozbalí PŮVODNÍ .xlsm; UPRAVENÝ .xlsm jen uloží (bez rozbalení)")
    args = parser.parse_args()

    log_path = Path(args.log).resolve()
    logger = setup_logger(log_path)

    input_xlsm = Path(args.input).resolve()
    if not input_xlsm.exists():
        logger.error(f"Soubor nenalezen: {input_xlsm}")
        write_run_separator(logger, log_path)
        sys.exit(EXIT_INPUT_NOT_FOUND)

    base_name = input_xlsm.stem

    try:
        # 1) Upravit ochrany v paměti (sheet + workbook)
        mod_bytes, stats = build_modified_xlsm_bytes(input_xlsm, args.sheet_xml, logger)

        # 2) Debug režim: rozbal PŮVODNÍ, ulož UPRAVENÝ (bez rozbalení)
        if args.debug:
            temp_dir = debug_temp_layout(input_xlsm, mod_bytes, base_name, args.sheet_xml, logger)
            logger.info(f"[DEBUG] TEMP složka: {temp_dir}")

        # 3) Načti data z listu přímo z bajtů upraveného .xlsm
        rows = load_sheet_rows_from_bytes(mod_bytes, args.sheet_name, logger)

        # 4) Vyřeš sloupce + názvy hlaviček (hlavička je vždy součástí výstupu)
        selected_indices, column_names, data_rows = resolve_columns(rows, args.cols, logger)
        logger.info(f"Vybrané indexy: {selected_indices}")
        logger.info(f"Názvy sloupců: {column_names}")
        logger.info(f"Počet datových řádků: {len(data_rows)}")

        # 5) Ulož výstupní .xlsx – header=True díky column_names
        df = build_filtered_dataframe(data_rows, selected_indices, column_names)
        out_path = compute_output_path(input_xlsm, args.out, args.sheet_name)

        try:
            df.to_excel(out_path, index=False, header=True, engine="openpyxl")
            logger.info(f"[OK] Uloženo: {out_path}")
        except Exception as e:
            logger.error(f"Zápis výstupního XLSX selhal: {e}")
            write_run_separator(logger, log_path)
            sys.exit(EXIT_WRITE_FAIL)

        write_run_separator(logger, log_path)
        sys.exit(EXIT_OK)

    except ValueError as e:
        # Specifické chyby (list nenalezen/prázdný/specifikace sloupců)
        msg = str(e)
        if msg in ("SHEET_NOT_FOUND", "SHEET_EMPTY"):
            write_run_separator(logger, log_path)
            sys.exit(EXIT_SHEET_ERROR)
        elif msg in ("COLSPEC_EMPTY", "COLSPEC_BAD_INDEX", "COLSPEC_HEADER_MISSING"):
            write_run_separator(logger, log_path)
            sys.exit(EXIT_COLSPEC_ERROR)
        else:
            logger.error(f"Neznámá ValueError: {e}")
            write_run_separator(logger, log_path)
            sys.exit(EXIT_UNKNOWN_ERROR)

    except Exception as e:
        logger.error(f"Chyba běhu: {e}")
        write_run_separator(logger, log_path)
        sys.exit(EXIT_INMEM_MOD_FAIL)


if __name__ == "__main__":
    main()
